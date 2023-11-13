import socket
import time
import openpyxl
import datetime

# Configuration
# HOST = '172.20.10.7'
HOST = '192.168.254.7'  # Listen to interfaces
PORT = 8888           # Port to listen on
BUFFER_SIZE = 1024 * 10  # Fixed buffer size 10KB

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active


def save_to_excel(row_data, row_num):
    # Convert the raw bytes to a string of hexadecimal values
    hex_data = ' '.join(f'{byte:02x}' for byte in row_data)

    # Get current time
    # current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Write binary data to the specified row
    ws.cell(row=row_num, column=1).value = hex_data
    # ws.cell(row=row_num, column=2).value = current_time
    wb.save('DMA_inter_data_mode2.xlsx')


def main():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind((HOST, PORT))
        s.listen()
        print(f"Listening on {HOST}:{PORT}...")

        row_num = 1
        while True:  # Infinite loop to keep the server running
            conn, addr = s.accept()
            with conn:
                print(f"Connected by {addr}")
                data = conn.recv(BUFFER_SIZE)
                # Write raw data to excel
                save_to_excel(data, row_num)
                row_num += 1

            #print(f"Connection from {addr} closed.")

                # print("Received:", data.decode('utf-8', errors='ignore'))



if __name__ == "__main__":
    main()


if __name__ == '__main__':
    print('Work done!')
